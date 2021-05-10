import openpyxl as xl
wb=xl.load_workbook('Book1.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
print(sheet.max_row)

for row in range(2, sheet.max_row+1):
    cell=sheet.cell(row, 3)
    corrected_value=cell.value*0.9
    corrected_price_cell=sheet.cell(row, 4)